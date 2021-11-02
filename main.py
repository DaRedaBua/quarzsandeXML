from log import *
import Bimail
from colorama import init, Fore, Style
import os
import datetime
import shutil
import re
import xlrd
import openpyxl
from tkinter.filedialog import askopenfilename, asksaveasfilename, askdirectory

# TODO: Afragen ob Aufträge erstellt werden sollen oder nicht!

# TODO: NiceToHave - Summe mit minusfahrten Franz schicken...

# debugging/logging
stacktrace = ""
recipients = []

# input Data
edvbookPath = ""
saveintern = ""
abrFilePath = ""
savespotPath = ""
extbookPath = ""

# XML
XMLerrOutPath = ""
XMLoutPath = ""

# UserDefinedVariables
dz = 0          #Dieselzuschlag
curMonth = 0    #Leistungsmonat
curYear = 0     #Leistungsjahr

# eingelesenes edvBook (xlrd book obj)
edvBook = xlrd.book.Book

# getZonePrices()
# Zonen, die im edvBook vorkommen
allowedZones = []
# dict zum nachschauen der Preise anhand der Zone (char)
zonePrices = {}

# getStrundenpreise() - dict zum nachschauen der Stundenpreise anhand Gerät (2A, 3A, 4A, etc..)
hourPrices = {}

# readExemptions()
# dict zum nachschauen der Ausnahmen - erklärung unten
Exemptions = {}
# Ein großes dict: key = aktion = so kann man später alle [preisprotonnen/keinTZ/keineMindermenge/...]berechnungen iterieren
#                  value = [Array] aller [Preisprotonnen/keinTZ/keineMindermenge]-Regeln
#                           Value/Array-Elem = Regel: Dict, die diese Regel beschreibt:
#                           regelDict: name: name,
#                                      action: entweder leer (keinTZ) oder Preis - je nach regelTyp
#                                      trigger: Array aller auslöser - ebenfalls array im [key,val]-Format, ALLE müssen true sein
#                                               ['zone','O'],['kz','ll123ab'],['kunde','leitl'],...
# ExemptionsDict
#   {'PreisProTonne': [
#            {  'action' : 3.5
#               'name'   : TreulAMW
#               'trigger': [['zone','O'],
#                           ['kunden','AMW']]
#            },{ 'action': 0.82, name: LeitlSand, triggger: [[kunden, leitl],[art,WE],[art,Sand]]},{...},{...},
#        ],
#   'PreisProStunde': [
#           {  'action'  : 90,4, name: Leidinger, trigger: [[kz, gr317cx]] },
#           {  'action'  : 99,7, name: Hope8, trigger: [[kz, uuhope8]] }
#        ]
#   }
# Unter wieviel Tonne zählt Mindermenge - wird deingelesen edv_abrechnung.xls
mindermenge = 0
# Wie viele Säcke pro tonne - wird eingelesen edv_abrechnung.xls
sackTonne = 0

# extractSubLicensePlates()
# Alle externen nicht-Reder kennzeichen zum schnellen nachschauen ob KZ überhaupt ext ist   x in extKZ
extKZ = []
# Zu jedem Kennzeichen wird (sofern verfügbar) der Besitzer gemappt  [ll123ab]:"reder"
extKZZuordnung = {}


# readAbrSheet()
KZOrders = {}
#   KZOrders['LL123AB']
#       [
#           dataDict - nur name, direkte adressierung der daten über [arr-index]
#               ['geraet]: F-3Achser
#               ['LFS-Datum']: 03.05.2021
#               ...,
#           dataDict,
#           dataDict,...
#      ]
#   KZOrders['LL123AB'][0]['geraet'] = F-3Achser
#   fehler-entry: [row,col,code] 0: Ausgelassene Int Fahrt, 1: zwei KZ, 2: Ausgelassene subfahrt 3: Minusgeschäft
fehler = []
# einglesene Grüne Liste (xlrd sheet obj)
srcsheet = xlrd.sheet.Sheet
# Anzahl der ausgelassenen Zeilen beim Einlesen
notCalculated = 0
# Kompletter DZ aus grüner Liste - wird gespeichert, weil später DZ pro Auto berechnet wird und dann soll == sein.
allIntDZ = 0
# String array mit 14 stellen
header = [""]*14

# extractSubfahrten()
# Anzahl an nicht berechneten Subfrächterfahrten
extNotCalculated = 0
# dict für alle gültigen Subfrächterfahrten
extOrders = {}

# calculateInterns()
# Berechnete Interne Fahrten - wird inital aus KZOrders kopiert und dann erweitert - gleicher aufbau wie KZOrders
KZOrdersCalc = {}
# Berechneter kompletter DZ - muss gleich/ähnlich sein mit allIntDZ aus grüner Liste
allIntCalcDZ = 0


# calculateSubContractors()
# dict für alle berechneten Subfrächterfahrten
extOrdersCalc = {}
# minus-Wert
minus = 0


def main():
    loadLogs("logs.csv")
    loadConfig()

    getZonePreise()
    getStundenPreise()
    readExemptionos()
    extractSubLicensePLates()

    askUser(1)

    readAbrSheet()
    extractSubFahrten()

    calculateInterns()
    calculateSubcontractors()

    consoleOut()
    minusOut()

def loadConfig():
    log(2000, "loadConfig", "")

    global recipients
    global outFolder
    global errFolder

    global edvbookPath
    global saveintern
    global abrFilePath
    global savespotPath
    global extbookPath

    global edvBook

    with open('config.csv', 'r') as f:
        lines = f.readlines()
        print(lines)
        for x in range(len(lines)):
            lineElems = lines[x].split(';')

            if x == 0:
                edvbookPath = lineElems[1]
                log(2005, "edvbookPath ", edvbookPath)

            if x == 1:
                saveintern = lineElems[1]
                log(2010, "savintern ", saveintern)

            if x == 2:
                abrFilePath = lineElems[1]
                log(2015, "abrFilePath ", abrFilePath)

            if x == 3:
                savespotPath = lineElems[1]
                log(2020, "savespotPath ", savespotPath)

            if x == 4:
                extbookPath = lineElems[1]
                log(2025, "extbookPath ", extbookPath)

            if x == 5:
                for e in range(1,len(lineElems)):
                    recipients.append(lineElems[e])
                log(2030, "recipients ", recipients)

    edvBook = xlrd.open_workbook(edvbookPath)


# Befüllt dict zonePrices mit zonenpreisen
# TODO: log
def getZonePreise():
    global edvBook

    edvSheet = edvBook.sheet_by_index(0)

    for i in range(2, 24):
        zone = edvSheet.cell(i, 0).value
        if isinstance(zone, float):
            zone = int(zone)
        zone = str(zone)
        allowedZones.append(zone)

        prices = [None]*3
        lastprice = 0;
        for y in range(0, 3):

            price = edvSheet.cell(i, 3-y).value
            if price != '':
                lastprice = price
            else:
                price = lastprice

            prices[2 - y] = price

        zonePrices[zone] = prices

    for i in range(27, 36):
        zone = edvSheet.cell(i, 0).value
        zone = str(zone)

        allowedZones.append(zone)

        price = float(edvSheet.cell(i, 1).value)
        prices = [price]*3

        zonePrices[zone] = prices


# befüllt dict hourPrices mit Stundenpreisen
# TODO: log
def getStundenPreise():
    edvSheet = edvBook.sheet_by_index(1)

    hourPrices["2A"] = float(edvSheet.cell(1, 0).value)
    hourPrices["3A"] = float(edvSheet.cell(1, 1).value)
    hourPrices["4A"] = float(edvSheet.cell(1, 2).value)
    hourPrices["Bagger"] = float(edvSheet.cell(1, 2).value)
    hourPrices["4A+"] = float(edvSheet.cell(1, 3).value)
    hourPrices["5A"] = float(edvSheet.cell(1, 4).value)
    hourPrices["HZ"] = float(edvSheet.cell(1, 4).value)
    hourPrices["Kran"] = float(edvSheet.cell(1, 5).value)


# befüllt Dict Exemptions aus DERZEIT edv_abrechnung.xls
#
# Ein großes dict:
#     key = aktion = so kann man später alle [preisprotonnen/keinTZ/keineMindermenge/...]berechnungen iterieren
#   value = [Array] aller [Preisprotonnen/keinTZ/keineMindermenge]-Regeln
#           [Value/Array-Elem] = Regel: Dict, die diese Regel beschreibt:
#               regelDict: name: name,
#               action: entweder leer (keinTZ) oder Preis - je nach regelTyp
#               trigger: [Array] aller auslöser - ebenfalls array im [key,val]-Format, ALLE müssen true sein
#                        ['zone','O'],['kz','ll123ab'],['kunde','leitl'],...
# ExemptionsDict
#   {'PreisProTonne': [
#            {  'action' : 3.5
#               'name'   : TreulAMW
#               'trigger': [['zone','O'],
#                           ['kunden','AMW']]
#            },{ 'action': 0.82, name: LeitlSand, triggger: [[kunden, leitl],[art,WE],[art,Sand]]},{...},{...},
#        ],
#   'PreisProStunde': [
#           {  'action'  : 90,4, name: Leidinger, trigger: [[kz, gr317cx]] },
#           {  'action'  : 99,7, name: Hope8, trigger: [[kz, uuhope8]] }
#        ]
#   }
def readExemptionos():
    log(6001, "readExemptions() aus Seite 3 - ", edvbookPath)
    edvSheet = edvBook.sheet_by_index(2)

    global Exemptions
    global mindermenge
    global sackTonne

    sackTonne = int(edvSheet.cell(0,1).value)
    log(6005, "var sackTonne: ", sackTonne)
    mindermenge = int(edvSheet.cell(1,1).value)
    log(6010, "mindermenge: ", mindermenge)

    # verschachteltes auslesen der xls - zuerst linke gruppierungsspalte
    for r in range(6, edvSheet.nrows):
        name = edvSheet.cell(r, 0).value
        type = edvSheet.cell(r, 1).value
        action = edvSheet.cell(r, 2).value

        log(6015, "name; aktion; parameter; ", str(name) + "; " + str(type) + "; " + str(action))

        # Wenn Regel namen und typ hat - parameter nur bei fixpreisen nötig
        if type != '' and type is not None and name != '' and name is not None:
            log(6020, "regel gültig - hat namen und typ", "")
            thisRule = {'action': action, 'trigger': [], 'name': name}

            # lege neue Ausnahme an, wenn noch nicht vorhandne
            if type not in Exemptions:
                log(6025, "lege neue Ausnahme an: ", name)
                Exemptions[type] = []

            # iteriere alle zeilen
            for rr in range(6, edvSheet.nrows):
                # suche trigger zu aktuellem namen
                if name == edvSheet.cell(rr, 4).value:
                    trigger = edvSheet.cell(rr, 5).value
                    value = edvSheet.cell(rr, 6).value
                    if isinstance(value, float):
                        value = str(int(value))
                    thisRule['trigger'].append([trigger, value])
                    log(6030, "Füge Trigger zu Regel hinzu: ", [trigger, value])

            #Füge sortiere Regel anhand des typs ein
            Exemptions[type].append(thisRule)


# liest Kennzeichen aus Frächteraufstellung.xls aus
# trägt alle nicht-reder-LKW in array extKZ ein - leichter zu prüfen >ob< kz extern ist
# trägt alle LKW (auch reder) in lookup dict ein - leichter zu prüfen, wem kz gehört: [uuhop8] => "leidinger"
# TODO: log
def extractSubLicensePLates():

    global extbookPath

    extBook = xlrd.open_workbook(extbookPath)
    extSheet = extBook.sheet_by_index(0)

    notpattern = re.compile('[a-z]')
    findpattern = re.compile('([A-Z]|[0-9]){3,9}')

    curName = ""

    for col in range(0, extSheet.ncols):
        for row in range(0, extSheet.nrows):
            val = extSheet.cell(row, col).value

            if row == 0:
                curName = val

            val = val.replace(' ','')
            val = val.replace('_4A','')
            vals = val.split('-')
            for part in vals:
                nono = notpattern.search(part)
                if nono is None:
                    yesyes = findpattern.search(part)
                    if yesyes is not None and len(part) < 10:
                        if part not in extKZ and curName != "Reder":
                            extKZ.append(part)
                        if part not in extKZZuordnung:
                            extKZZuordnung[part] = curName


def askUser(key):
    log(3000, "askUser(key)", key)

    global dz
    global curYear
    global curMonth
    global srcsheet

    if key == 1:
        log(3001, "askUser() - 1", "")
        print(Fore.CYAN + Style.BRIGHT + "Willkommen im Quarzsande Programm v2.0!")

        print(Fore.CYAN + Style.BRIGHT + "Geben Sie bitte den Treibstoffzuschlag in % ein! (Bei Subfrächtergutschriften wird automatisch +2% Aufschlag gerechnet!)")
        dp = input()
        dp = dp.replace("%", "")
        dp = dp.replace(",", ".")
        dp = float(dp)
        dp /= 100
        dp += 1
        dz = dp

        log(3005, "dz: ", dz)

        while curMonth <= 0 or curMonth > 12 or curYear < 21 or curYear > 50:
            print(Fore.CYAN + Style.BRIGHT + "\nGeben Sie den Leistungszeitraum ein! mm/jj - z.B.: '09/21'")
            inp = input()
            inp = inp.split("/")
            curMonth = int(inp[0])
            curYear = int(inp[1])
            log(3010, "curMonth: ", curMonth)
            log(3011, "curYear: ", curYear)

        print(Fore.CYAN + Style.BRIGHT + "Drücken Sie Enter, um die Fahrten-Datei auszuwählen")
        input()
        abrFile = askopenfilename(initialdir=abrFilePath)
        abrBook = xlrd.open_workbook(abrFile)
        srcsheet = abrBook.sheet_by_index(0)

        log(3015, "Grüne Liste, abrFile: ", abrFile)


def createFolders():
    log(4000, "createFolders()", "")
    global curYear
    global curMonth

    indir = saveintern + "20" + curYear + "_" + curMonth
    exdir = saveintern + "Subfrächter_20" + curYear + "_" + curMonth

    log(4005, "Neuer Interner ouput-Ordner, indir: ", indir)
    log(4010, "Neuer Subfrächter Ausgabe-Ordner: ", exdir)

    if os.path.exists(exdir):
        shutil.rmtree(exdir)

    if os.path.exists(indir):
        shutil.rmtree(indir)

    os.mkdir(exdir)
    os.mkdir(indir)


def extractLicensePlates(orplate):
    log(5000, "extractLicensePlate(orplate); orplate: ", orplate)

    entr = orplate.replace('-','')
    entr = entr.replace('_4A','')

    if orplate == '':
        entr = "KEIN"

    log(5005, "entr für verarbeitung: ", entr)

    notpattern = re.compile('[a-z]')
    findpattern = re.compile('([A-Z]|[0-9]){3,9}')

    platesFound = []

    # TODO: besser machen - derzeit könnte ABCD ein kennzeichen sein aber ll12md nicht
    # TODO: trim
    # TODO: edgeCases ausprobieren...

    plates = entr.split('+')       # Zuerst wird am + gepslittet
    for plate in plates:
        parts = plate.split(',')    # Dann an einem Beistrich - muss man hintereinander machen
        for part in parts:
            nono = notpattern.search(part)  # Ausschlussverfahren - wenn nur buchstaben => ungültig
            if nono is None:
                yesyes = findpattern.search(part)   # Suchverfahren - Muss Großbuchstaben
                if yesyes is not None and len(part) < 10:
                    platesFound.append(part)

    log(5030, "platesFound: ", platesFound)

    return platesFound


# TODO: STEHEN LASSEN - WENN SICH EXCEL-LISTE VERSCHIEBT HIER ÄNDERN!
# readAbrSheet teilt Zeilen nach Kennzeichen auf.
# Als Ergebnis dieser Methode ist das Dictionary KZOrders befüllt. Pro dict-key ein array mit dict pro fahrt.
# KZOrders['LL123AB']
#   [
#       dataDict (data ist nur hier der variablen-Name für das Dict)
#           ['geraet]:F-3Achser
#           ['LFS-Datum']:F-Achser
#           ...,
#       dataDict,
#       dataDict,...
#   ]
# Fehler werden nicht als KZ-Orders-Einträge behandelt sondern als hinweise auf originale Zelle in grüner Excel-Liste
#   (Zeile, Spalte, Fehlercode 0/1/2/3)
def readAbrSheet():
    log(6000, "readAbrSheet(srcsheet); srcsheet: ", srcsheet)

    global notCalculated
    global allIntDZ
    global header
    global fehler
    global KZOrders

    # Sammelt Überschriften aus Zeile 10 und DANCH Zeile 9 - hat nix mit einem Auftrag speziell zu tun
    # Notwendig, weil Gerät, LFS-nr, kunde, etc. in Zeile 10 in B-K steht (siehe obere For-Loop)
    #    Menge, Stunden, Ger.Kosten, Mautkosten aber in Zeile 8 L-O
    #    Untere Loop überschreibt leere Zellen 10 L-O mit gefüllten Zellen 9 L-O
    for c in range(1, srcsheet.ncols):
        header[c-1] = srcsheet.cell(9, c).value
    for c in range(srcsheet.ncols-4, srcsheet.ncols):
        header[c-1] = srcsheet.cell(8, c).value

    log(6005, "header: ", header)

    # Sortiert Zeile für Zeile bei Kennzeichen ein
    for i in range(10, srcsheet.nrows):
        success = True

        # In data wird eine Zeile/Fahrt als Dict gespeichert.
        data = dict()
        data['geraet'] = srcsheet.cell(i, 1).value
        data['lfs_datum'] = srcsheet.cell(i, 2).value
        data['lfs_nr'] = srcsheet.cell(i, 3).value
        data['art_lfrnt'] = srcsheet.cell(i, 4).value
        data['art'] = srcsheet.cell(i, 5).value
        data['kunden'] = srcsheet.cell(i, 6).value
        data['baustelle'] = srcsheet.cell(i, 7).value
        data['zone'] = srcsheet.cell(i, 9).value
        data['einheit'] = srcsheet.cell(i, 10).value
        data['menge'] = srcsheet.cell(i, 11).value
        data['stunden'] = srcsheet.cell(i, 12).value
        data['ger_kosten'] = srcsheet.cell(i, 13).value
        data['mautk'] = srcsheet.cell(i, 14).value
        data['anmerkungen'] = ""
        data['zeile'] = i+1

        # Zone (alphanumerisch) fix in UPPER String umwandeln
        if isinstance(data['zone'], str):
            data['zone'] = data['zone'].upper()
        else:
            data['zone'] = str(int(data['zone']))

        # Herausfinden der Kennzeichen in dieser Zeile, gibt array zurück, [plates]
        kz = srcsheet.cell(i, 8).value
        plates = extractLicensePlates(kz)

        # doFehler in verbindung mit Success:
        # Weiter unten überprüft mein Programm die einzelnen Zeilen auf verarbeitbarkeit
        # Wenn Fehler passieren, wird die glob Variable fehler appendet.
        # Allerdings sollen die Zeilen Treibstoff und die Summenzeile das nicht auslösen.
        # Quickfix: naturally true until proven false, siehe nächste Zwei IFs
        doFehler = True

        # Oberer Kommentar gilt, Bei Treibstoffzeile wird Wert gemerkt um später mit Summen der einzelautos zu vgl.
        if "Treibstoff" in kz:
            allIntDZ = data['ger_kosten']
            doFehler = False

        if "Summe" in data['geraet']:
            doFehler = False


        # Wenn mehr als 1 Kennzeichen in Zeile -> Leichter Error (wird im moment nicht behandelt)
        if(len(plates) > 1) and doFehler:
            fehler.append([i, 7, 1])

        # Kein Kennzeichen => Schwerer Fehler
        if(len(plates) == 0) and doFehler:
            success = False
            fehler.append([i, 7, 0])

        # Keine Gerätekosten
        if data['ger_kosten'] is None or data['ger_kosten'] == '' or data['ger_kosten'] == 0 and doFehler:
            fehler.append([i, 12, 0])
            success = False

        # Keine Lieferscheinnummer
        if data['lfs_nr'] is None or data['lfs_nr'] == '' or data['lfs_nr'] == 0:
            fehler.append([i, 2, 0])
            success = False

        # Wenn Treibstoff oder Summenzeile => nicht berechnen (suc false), aber keinen Fehler schreiben
        if not doFehler:
            success = False

        # Wenn Zeile soweit ok, Fahrt im Kennzeichen-Dictionary einordnen
        if success:
            if plates[0] not in KZOrders:   # Wenn Kennzeichen noch nicht im Dictionary => Anlegen
                KZOrders[plates[0]] = []
            data['kz'] = plates[0]          # Erst hier, weil certain len(plates) > 0
            KZOrders[plates[0]].append(data)    # data-Obj wird hinzugefügt
            log(6050, "Kennzeichen: ", data['kz'])
            log(6051, "data: ", data)
        elif doFehler:
            notCalculated += 1              # Summe an ausgelassenen Datensätzen wird um 1 erhöht



def extractSubFahrten():
    log(10000, "extractSubFahrten()", "")

    global extNotCalculated
    global extOrders

    log(10005, "iterate over all extKZ: ", extKZ)
    for kz in extKZ:
        log(10010, "kz: ", kz)
        if kz in KZOrders and extKZZuordnung[kz] != "Reder":
            log(10015, "kz has Orders and does not belong to Reder - iterate over KZOrders[kz]", KZOrders[kz])
            orders = []

            for fahrt in KZOrders[kz]:

                log(10020, "fahrt = elem aus KZOrders: ", fahrt)

                # Naturally true until proven false
                success = True

                # Todo: ALTER CODE
                # found = False
                # for z in allowedZones:
                #    if fahrt['zone'] == z:
                #        found = True
                # if fahrt['zone'] is None or fahrt['zone'] == '' or not found:

                # Wenn Zone leer oder Zone nicht in erlaubten Zonen und keine Pauschalfahrt => Subfrächterfehler
                if fahrt['zone'] is None or fahrt['zone'] == '' or fahrt['zone'] not in allowedZones:
                    if fahrt['einheit'] != 'pau':
                        fehler.append([fahrt['zeile'], 8, 2])
                        success = False

                # Wenn einheit nicht vorhanden oder nicht eindeutig => subfrächterfehler
                if fahrt['einheit'] == '' or fahrt['einheit'] is None or fahrt['einheit'] == 'stk':
                    fehler.append([fahrt['zeile'], 9, 2])
                    success = False

                # Wenn menge nicht vorhanden, keine pauschalfahrt und stunden nicht vorhanden => Subfrächterfehler
                if fahrt['menge'] == '' or fahrt['menge'] == 0 or fahrt['menge'] is None:
                    if fahrt['einheit'] != 'pau':
                        if fahrt['stunden'] == '' or fahrt['stunden'] == 0 or fahrt['stunden'] is None:
                            fehler.append([fahrt['zeile'], 10, 2])
                            fehler.append([fahrt['zeile'], 11, 2])
                            success = False

                # Wenn kein gerät und keine Pauschalfahrt => Subfrächterfehler
                if fahrt['geraet'] == '' or fahrt['geraet'] == 0 or fahrt['geraet'] is None:
                    if fahrt['einheit'] != 'pau':
                        fehler.append([fahrt['zeile'], 0, 2])
                        success = False

                # Wenn gültige fahrt
                if success:
                    fahrt['orig_kosten'] = (fahrt['ger_kosten']*dz) + fahrt['mautk']
                    orders.append(fahrt)
                    log(10025, "fahrt gültig für Subfrächterberechnung- siehe oben - zeile: ", fahrt['zeile'])
                else:
                    extNotCalculated += 1
                    log(10030, "fahrt ungültig für Subfrächterberechnung - siehe oben - zeile: ", fahrt['zeile'])

            extOrders[kz] = orders


# Berechnet pro Fahrt Dieselzuschlag und Gesamtkosten für diese Fahrt.
# allIntCalcDZ wird pro Fahrt erhöht und soll dann gleich sein mit Treibstoffzeilenwert aus grüner Liste
# TODO: log
def calculateInterns():

    global KZOrdersCalc
    KZOrdersCalc = KZOrders

    global allIntCalcDZ

    for kz, orders in KZOrdersCalc.items():
        for fahrt in orders:
            fahrt['dz_abs'] = fahrt['ger_kosten'] * (dz - 1)
            allIntCalcDZ += fahrt['dz_abs']
            fahrt['summe'] = (fahrt['ger_kosten'] * dz) + fahrt['mautk']


# TODO: log
def calculateSubcontractors():
    global extOrdersCalc
    global minus


    # Liste wird zuerst kopiert und dann wird IN den einzelnen Elems die Berechnung durchgeführt.
    extOrdersCalc = extOrders

    for kz, orders in extOrdersCalc.items():
        for fahrt in orders:

            # Mindermenge und DZ werden standardmäßig berechnet
            doMind = True
            dodz = True;

        # Sack ausrechnen - keine Mindermenge bei Sack
            if fahrt['einheit'] == 'Sack':
                fahrt['anmerkungen'] = str(int(fahrt['menge'])) + " Säcke geladen. 40 Säcke pro Tonne laut Tarifblatt"
                fahrt['menge'] = fahrt['menge']/40
                fahrt['einheit'] = 'to'
                doMind = False

        # Mindermengem-Ausnahme aus EDV_Abrechnung.xls
        # 500 IQ code...Kommentare lesen, dann verstehen!
            # Normalerweise, führe Mindermengenberechnung durch, default true, wird weiter unten bei # XX benötigt.
            do = True
            # iteriere alle einzelnen Mindermengen-Regeln
            for rule in Exemptions['KeineMindermenge']:
                # Nimmt sich die Trigger Anzahl dieser Regel - zb 3
                triggersToGo = len(rule['trigger'])
                # Geht durch jeden trigger
                for specifier in rule['trigger']:
                    # Schaut ob Trigger auslöst (trigger 'zone','U' -> if 'U' in fahrt['zone'] => T/F
                    if specifier[1] in fahrt[specifier[0]]:
                        # Wenn trigger true -> reduziere um 1 (zb 3 -> 2), gehe nächstes trigger durch
                        triggersToGo -= 1
                # NUR WENN ALLE TRIGGER AUSLÖSEN GEHT ES RICHTUNG null ... 3 -> 2 -> 1 -> 0.
                if triggersToGo == 0:
                    # Dann setze do auf false, weil wir wollen ja KEINE MINDERMENGE berechnung, for schleife kann quit
                    do = False
                    break
            # Oberer Code entscheided, ob hier die mindermengenregel (zb 12to) angewendet wird.
            if do and fahrt['einheit'] == 'to':
                if fahrt['menge'] < mindermenge and doMind:
                    fahrt['anmerkungen'] = str(fahrt['menge']) + "to - Mindermenge - 12to gerechnet."
                    fahrt['menge'] = mindermenge


            # Bei pauschalbeträgen gibt es keinen DZ, wenn keine pau fahrt, dann setze lookup zeichen 2A, 3A, 4A, 5A
            if fahrt['einheit'] == 'pau':
                dodz = False
            else:
                #if kz.lower() == "uuhope3":
                #    if fahrt['lfs_nr'] == 501714:
                #        print("hallo1")
                #    if fahrt['zone'] == 4 or fahrt['zone'] == '4':
                #        print("hallo1")
                #    print("heraussen")

                if "2Achs" in fahrt['geraet']:
                    z = 0
                    st = '2A'
                elif "3Achs" in fahrt['geraet']:
                    z = 0
                    st = '3A'
                elif "4Achs" in fahrt['geraet']:
                    z = 1
                    st = '4A'
                elif "5Achs" in fahrt['geraet']:
                    z = 2
                    st = '5A'
                else:
                    success = False

            #Tonnage Preise rechnen + Ausnahmen
            if fahrt['einheit'] == 'to':
                newPrice = False
                for rule in Exemptions['PreisProTonne']:
                    isZero = len(rule['trigger'])
                    for specifier in rule['trigger']:
                        if specifier[1] in fahrt[specifier[0]]:
                            isZero -= 1
                    if isZero == 0:
                        newPrice = float(rule['action'])
                        break
                if not newPrice:
                    preis = zonePrices[fahrt['zone']][z]
                    fahrt['ger_kosten'] = float(fahrt['menge'])*preis
                elif newPrice:
                    fahrt['ger_kosten'] = float(fahrt['menge'])*newPrice

            # Stunden Preise rechnen + Ausnahmen
            if fahrt['einheit'] == 'std':
                newPrice = False
                for rule in Exemptions['PreisProStunde']:
                    isZero = len(rule['trigger'])
                    for specifier in rule['trigger']:
                        if specifier[1] in fahrt[specifier[0]]:
                            isZero -= 1
                    if isZero == 0:
                        newPrice = float(rule['action'])
                        break
                if not newPrice:
                    preis = hourPrices[st]
                else:
                    preis = newPrice

                if fahrt['stunden'] == '' or fahrt['stunden'] is None:
                    fahrt['ger_kosten'] = float(fahrt['menge'])*preis
                else:
                    fahrt['ger_kosten'] = float(fahrt['stunden'])*preis

            # Dieselzuschalg einrechnen + Ausnahme
            for rule in Exemptions['KeinTZ']:
                isZero = len(rule['trigger'])
                for specifier in rule['trigger']:
                    if specifier[1] in fahrt[specifier[0]]:
                        isZero -= 1
                if isZero == 0:
                    dodz = False
                    break
            DZ = dz
            if dodz:
                DZ += 0.02
            else:
                DZ = 1

            fahrt['dz_abs'] = fahrt['ger_kosten'] * (DZ-1)

            do = True
            for rule in Exemptions['KeineMaut']:
                isZero = len(rule['trigger'])
                for specifier in rule['trigger']:
                    if specifier[1] in fahrt[specifier[0]]:
                        isZero -= 1
                if isZero == 0:
                    do = False
                    break
            if not do and fahrt['mautk'] > 0:
                fahrt['mautk'] = 0
                fahrt['anmerkungen'] += "Maut wie vereinbart auf 0€ gesetzt"

            fahrt['summe'] = (fahrt['ger_kosten'] * DZ) + fahrt['mautk']

            #Machen wir ein Minusgschäft?
            if fahrt['summe'] > fahrt['orig_kosten']:
                minus += (fahrt['summe']-fahrt['orig_kosten'])
                fehler.append([fahrt['zeile'], 0, 3, fahrt])


def consoleOut():

    if allIntDZ == 0:
        print(Fore.RED + "Konnte den Treibstoffzuschlag nicht in der Liste finden. Wenn Sie die Differenz sehen möchten, geben Sie bitte Treibstoff im Feld Kennz. ein!\n")

    print( "TZ laut grüner Liste: ", Style.BRIGHT + str(allIntDZ))
    print("Errechneter TZ: ", Style.BRIGHT + str(round(allIntCalcDZ, 2)))
    print(Style.BRIGHT + "\tDifferenz: ", Fore.YELLOW + Style.BRIGHT + str(round(allIntDZ-allIntCalcDZ, 2)))
    print(Style.BRIGHT + "\nAusgelassene Datensätze - interne Verbuchung: ", Style.BRIGHT + str(notCalculated))
    for f in fehler:
        if f[2] == 0:
            print("\tZeile: ", Fore.YELLOW + str((f[0]+1)), "  -  " , Fore.YELLOW + header[f[1]])

    print(Style.BRIGHT + "\nAusgelassene Datensätze - Subfrächtergutschriften: ", Style.BRIGHT + str(extNotCalculated))

    for f in fehler:
        if f[2] == 2:
            print("\tZeile: ", Fore.YELLOW + str((f[0])), "  -  " , Fore.YELLOW + header[f[1]])

    print(Style.BRIGHT + "\nSumme aller Minus-Fahrten: ", Style.BRIGHT + Fore.RED + str(round(minus, 2)))


def minusOut():

    global savespotPath

    minusBook = openpyxl.Workbook()
    minusSheet = minusBook.active

    for h in range(14):
        minusSheet.cell(1, h+1).value = header[h]

    minusSheet.cell(1, 15).value = "Zeile (Grüne Liste)"
    minusSheet.cell(1, 13).value = "Eingang = Ger.Kosten + Maut + TZ"
    minusSheet.cell(1, 16).value = "Subfrächtergutschift"

    mapping = [1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 12, 13, 14, None, 15, 8, None, 16, None, None]

    r = 1
    for f in fehler:
        if f[2] == 3:
            r += 1
            c = 0
            for key, val in f[3].items():
                if c <=21:
                    if mapping[c] is not None:
                        minusSheet.cell(r, mapping[c]).value = val
                    c += 1

    widths = [
        ['A', 15],
        ['B', 10],
        ['C', 10],
        ['D', 18],
        ['E', 25],
        ['F', 35],
        ['G', 25],
        ['H', 10],
        ['I', 10],
        ['J', 9],
        ['K', 9],
        ['L', 10],
        ['M', 35],
        ['N', 15],
        ['O', 20],
        ['P', 20],
        ['Q', 20],
        ['R', 20],
        ['S', 20],
        ['T', 20],
        ['U', 20]
    ]

    for i in range(0, minusSheet.max_column, 1):
        minusSheet.column_dimensions[widths[i][0]].width = widths[i][1]

    print(Fore.CYAN + Style.BRIGHT + "\nGeben Sie bitte an, wo die Minusfahrten-Liste abgespeichert werden soll. Drücken Sie ENTER zum auswählen.")
    input()

    #TODO; return!
    saveSpot = asksaveasfilename(initialdir=savespotPath, defaultextension=".xlsx", initialfile="Minusfahrten_" + str(curYear) + '_' + str(curMonth))

    minusBook.save(saveSpot)


main()